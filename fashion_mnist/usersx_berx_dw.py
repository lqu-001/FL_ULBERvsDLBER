
from openpyxl import Workbook #xls
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import copy
import numpy as np
np.set_printoptions(threshold=np.inf)
from torchvision import datasets, transforms
import torch
import math
import random

from utils.sampling import mnist_iid, mnist_noniid, fashionmnist_iid, fashionmnist_noniid, cifar_iid
from utils.options import args_parser
from models.Update import LocalUpdate
from models.Nets import mlp, lenet_300_100, lenet5, vanillacnn, cnn3, resnet18
from models.Fed import FedAvg
from models.mask import *
from models.test import test_img

import time
start=time.time()

args = args_parser()

seed=args.seed
np.random.seed(seed)
torch.manual_seed(seed)
torch.cuda.manual_seed_all(seed)
torch.backends.cudnn.deterministic = True
torch.backends.cudnn.benchmark = False

if __name__ == '__main__':    
    args.device = torch.device('cuda:{}'.format(args.gpu) if torch.cuda.is_available() and args.gpu != -1 else 'cpu')        
#dataset------------------------------------------------------------------------------------------------------
    if args.dataset == 'mnist':
        trans_mnist = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.1307,), (0.3081,))])
        dataset_train = datasets.MNIST('./data/mnist/', train=True, download=True, transform=trans_mnist)
        dataset_test = datasets.MNIST('./data/mnist/', train=False, download=True, transform=trans_mnist)
        if args.iid:
            dict_users = mnist_iid(dataset_train, args.num_users)
        else:
            dict_users = mnist_noniid(dataset_train, args.num_users)
    elif args.dataset == 'fashionmnist': 
        transform = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.5,), (0.5,))]) 
        dataset_train = datasets.FashionMNIST('./data/FashionMNIST', download=True, train=True, transform=transform)
        dataset_test = datasets.FashionMNIST('./data/FashionMNIST', download=True, train=False, transform=transform)
        if args.iid:
            dict_users = fashionmnist_iid(dataset_train, args.num_users)
        else:
            dict_users = fashionmnist_noniid(dataset_train, args.num_users)        
    elif args.dataset == 'cifar10':        
        transform1 = transforms.Compose([transforms.RandomHorizontalFlip(),transforms.RandomGrayscale(),transforms.ToTensor(),transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))])
        transform2 = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))])
        dataset_train = datasets.CIFAR10('./data/cifar', train=True, download=True, transform=transform1)
        dataset_test = datasets.CIFAR10('./data/cifar', train=False, download=True, transform=transform2)
        if args.iid:
            dict_users = cifar_iid(dataset_train, args.num_users)
        else:
            exit('Error: only consider IID setting in CIFAR10')
            #dict_users = cifar_noniid(dataset_train, args.num_users)
    else:
        exit('Error: unrecognized dataset')
    img_size = dataset_train[0][0].shape  
 
#model---------------------------------------------------------------    
    
    if args.model == 'mlp':
        net_server = mlp(args=args).to(args.device)
    elif args.model == 'lenet_300_100':
        net_server = lenet_300_100(args=args).to(args.device)
    elif args.model == 'lenet5':
        net_server = lenet5(args=args).to(args.device)
    elif args.model == 'vanillacnn':
        net_server = vanillacnn(args=args).to(args.device)
    elif args.model == 'cnn3':
        net_server = cnn3(args=args).to(args.device)
    elif args.model == 'resnet18':
        net_server = resnet18().to(args.device)
    else:
        exit('Error: unrecognized model')
    #print(net_server) 
    #net_server.train()    
    for key in net_server.state_dict().keys():
        print(key, net_server.state_dict()[key].size())
        
#server to clients-------------------------------------------    
    glob_acc, glob_train_loss = [],[]
    wb = Workbook() #xls
    ws = wb.active  #xls
    upbitsvol=0; dnbitsvol=0
#------------rounds------------------------------------------  
    for rounds in range(args.rounds):
        serverbroad = copy.deepcopy(net_server.state_dict())
#---------------------server broadcast----------------------- 
        wglob = copy.deepcopy(net_server.state_dict())
        tmp = copy.deepcopy(net_server.state_dict())
        wldpc = copy.deepcopy(net_server.state_dict())##
        net_client_in = copy.deepcopy(net_server.to(args.device))##        
        dw_locals=[]; loss_locals = []
        param_num = sum(wldpc[i].numel() for i in wldpc.keys()) #582026       
#----------------get fp32 1d-sequence: seqfp32---------------------------------        
        seqfp32 = wldpc['conv1.weight'].flatten() #flatten horizontal
        for key in wldpc.keys():
            if (key != 'conv1.weight'):
                seqfp32 = torch.cat((seqfp32, wldpc[key].flatten()), -1) #torch.tensor fp32
        seqbin, factor, mini = quan_bit(seqfp32, args.dnbit) #seqbin=['1001','0010'...],list
        for j in range (len(seqbin)):
            seqbin[j]=list(seqbin[j]) #list seqbin=[['1','0','0','1'],['0','0','1','0']...]       
#-------------choose flip location and get 1d-sequence: seqfp32_ber----------------------------------        
        #np.random.seed(4)
        location = np.random.choice(param_num*args.dnbit, int(param_num*args.dnbit*args.dnber), replace=False)#choose location to introduce ber
        for index in location:
            item = index//args.dnbit
            bit = index%args.dnbit
            if (seqbin[item][bit]=='1'):
                seqbin[item][bit]='0'
            else:
                seqbin[item][bit]='1'       #after ber, seqbin=[['0','0','0','1'],['0','0','1','0']...]
        for k in range (len(seqbin)):
            seqbin[k]=''.join(seqbin[k])    # seqbin=['0001','0010'...]
        seqfp32_ber = dequan_bit(seqbin,factor, mini, args.dnbit) #seqfp32_ber=[0.32,-0.02,...]
        seqfp32_ber = torch.Tensor(seqfp32_ber).to(args.device) #torch.tensor fp32        
#-----------rehape and load the noised param to NN------------------------------        
        coodi=0
        for p in wldpc.keys():
            coodi += wldpc[p].numel()
            layer_ber = (seqfp32_ber[coodi-wldpc[p].numel():coodi]).reshape(wldpc[p].shape)
            net_client_in.state_dict()[p].data.copy_(layer_ber)
        net_client_in.to(args.device)
        net_client_in.eval()             
#-----------local train-----------------------------------
        client_join = max(int(args.frac * args.num_users), 1)
        idxs_users = np.random.choice(args.num_users, client_join, replace=False) #[0,3,5,9,2,...]; np.random: equal problity; False means not repeat
        for idx in idxs_users:         
            local = LocalUpdate(args=args, dataset=dataset_train, idxs=dict_users[idx])
            w, mode, loss = local.train(net=copy.deepcopy(net_client_in).to(args.device))
            for k in w.keys():                
                dw=copy.deepcopy(w[k])-net_client_in.state_dict()[k]                   
                tmp[k].data.copy_(dw)
                
            ##client quant and ber----------------------------------------------------------
            seqfp32up = tmp['conv1.weight'].flatten() #flatten horizontal
            for key in tmp.keys():
                if (key != 'conv1.weight'):
                    seqfp32up = torch.cat((seqfp32up, tmp[key].flatten()), -1) #torch.tensor fp32
            seqbin, factor, mini = quan_bit(seqfp32up, args.upbit) #seqbin=['1001','0010'...],list
            for j in range (len(seqbin)):
                seqbin[j]=list(seqbin[j]) #list seqbin=[['1','0','0','1'],['0','0','1','0']...]       
            #choose flip location and get 1d-sequence: seqfp32up_ber----------------------------------        
            np.random.seed(rounds+idx)
            location = np.random.choice(param_num*args.upbit, int(param_num*args.upbit*args.upber), replace=False)#choose location to introduce ber
            for index in location:
                item = index//args.upbit
                bit = index%args.upbit
                if (seqbin[item][bit]=='1'):
                    seqbin[item][bit]='0'
                else:
                    seqbin[item][bit]='1'       #after ber, seqbin=[['0','0','0','1'],['0','0','1','0']...]
            for k in range (len(seqbin)):
                seqbin[k]=''.join(seqbin[k])    # seqbin=['0001','0010'...]
            seqfp32up_ber = dequan_bit(seqbin,factor, mini, args.upbit) #seqfp32up_ber=[0.32,-0.02,...]
            seqfp32up_ber = torch.Tensor(seqfp32up_ber).to(args.device) #torch.tensor fp32        
            #rehape and load the noised param to NN------------------------------        
            coodi=0
            for p in tmp.keys():
                coodi += tmp[p].numel()
                layer_ber = (seqfp32up_ber[coodi-tmp[p].numel():coodi]).reshape(tmp[p].shape)
                tmp[p].data.copy_(layer_ber)
            ##server receive after quant and ber----------------------------------------------------------
            
            dw_locals.append(copy.deepcopy(tmp))
            loss_locals.append(copy.deepcopy(loss))
        dw_glob=FedAvg(dw_locals)
#-----------server agg and train test---------------------        
        for k in wglob.keys():
            wglob[k]+=dw_glob[k]
        net_server.load_state_dict(wglob)
        
        net_server.to(args.device)
        net_server.eval()
        test_acc, test_loss = test_img(net_server, dataset_test, args)
        train_acc, train_loss = test_img(net_server, dataset_train, args) 
        avg_loss = sum(loss_locals) / len(loss_locals)
        
        print(test_acc.numpy())
        print('Avg_loss:',avg_loss)
        print('Train_loss {:.3f}, Test_loss {:.3f}'.format(train_loss, test_loss))
        
        upbitsvol += client_join*args.upbit*param_num/(10**9) #Gb
        dnbitsvol += client_join*args.dnbit*param_num/(10**9) #Gb
        
        ws.cell(rounds+2,1).value = rounds  
        ws.cell(rounds+2,2).value = str(test_acc.numpy()) 
        ws.cell(rounds+2,3).value = str(test_loss)
        ws.cell(rounds+2,4).value = str(train_acc.numpy())
        ws.cell(rounds+2,5).value = str(train_loss)
        ws.cell(rounds+2,6).value = str(avg_loss)
        ws.cell(rounds+2,7).value = upbitsvol
        ws.cell(rounds+2,8).value = dnbitsvol
                                    
    wb.save('./result/user'+str(args.num_users)+'upber'+str(args.upber)+'_dw.xlsx')
    
end=time.time()
print('running time is ',end-start)
